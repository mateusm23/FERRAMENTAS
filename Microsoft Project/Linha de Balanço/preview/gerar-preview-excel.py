# -*- coding: utf-8 -*-
#
# GERADOR DE PREVIEW DE EXCEL (protótipo em Python)
# ---------------------------------------------------
# Este script e o "codigo por tras" do arquivo export-excel-preview.xlsx que temos revisado.
# Ele roda so aqui, do meu lado (Python + openpyxl), pra prototipar rapido a formatacao visual
# com dado ficticio antes de decidir como fica a versao final.
#
# IMPORTANTE: a ferramenta real (gerador-linha-de-balanco.html) e 100% HTML/JS, sem backend
# nem Python. Quando formos implementar o export de verdade, essa mesma logica de layout
# (cabecalho, cores, raias, bordas) sera reescrita em JavaScript usando uma biblioteca como
# ExcelJS, rodando direto no navegador a partir da classificacao real ja carregada. Este
# script Python NAO faz parte do gerador-linha-de-balanco.html - e so uma ferramenta de
# rascunho, mantida aqui no preview/ pra voce poder ler e acompanhar a logica por tras
# de cada ajuste visual antes de portarmos pra versao definitiva em JS.
import datetime
import math
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MESES_PT = ["JANEIRO", "FEVEREIRO", "MARCO", "ABRIL", "MAIO", "JUNHO", "JULHO",
            "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]


def mes_label_pt(date):
    return f"{MESES_PT[date.month - 1]} DE {date.year}"


def texto_contraste(hex_cor):
    r, g, b = int(hex_cor[0:2], 16), int(hex_cor[2:4], 16), int(hex_cor[4:6], 16)
    luminancia = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return "1A2B45" if luminancia > 0.55 else "FFFFFF"


NAVY = "1A2B45"
AZUL_CONTRASTE = "1A6EE8"       # tarja fininha decorativa embaixo do banner
CINZA_PAV_COL = "EAF0F8"
CINZA_MES_HDR = "D3E3F5"
CINZA_FIMSEMANA = "EAEEF3"
CINZA_FIMSEMANA_HDR = "B8CBE0"
FILL_PAV_PAR = "F7FAFD"
MUTED = "6B80A0"
COR_DIA = "33455E"
BORDA_SUAVE_COR = "9AA9C0"
BORDA_MES_COR = "3D5A80"  # antes era o navy quase preto; um azul mais suave, ainda distinto


def borda_com(cel, **novos_lados):
    atual = cel.border
    base = {"left": atual.left, "right": atual.right, "top": atual.top, "bottom": atual.bottom}
    base.update(novos_lados)
    cel.border = Border(**base)


BORDA_PAV_SEP = Side(style="thin", color=BORDA_SUAVE_COR)
BORDA_MES = Side(style="medium", color=BORDA_MES_COR)
BORDA_DIA_CLARA = Side(style="thin", color="E3E8F0")
BORDA_DIA_HDR = Side(style="thin", color=BORDA_SUAVE_COR)
BORDA_FINA = Border(left=Side(style="thin", color="D4DFE8"), right=Side(style="thin", color="D4DFE8"),
                     top=Side(style="thin", color="D4DFE8"), bottom=Side(style="thin", color="D4DFE8"))

FONTE_DIA = Font(name="Calibri", bold=True, size=9, color=COR_DIA)
FONTE_DIA_FIMSEMANA = Font(name="Calibri", bold=True, size=9, color="C53030")
FONTE_PAV = Font(name="Calibri", bold=True, size=10, color=NAVY)
FILL_FIMSEMANA = PatternFill("solid", fgColor=CINZA_FIMSEMANA)
FILL_FIMSEMANA_HDR = PatternFill("solid", fgColor=CINZA_FIMSEMANA_HDR)
FILL_PAV_PAR_O = PatternFill("solid", fgColor=FILL_PAV_PAR)

SERVICOS = [
    ("Alvenaria", "0F9B6E"),
    ("Eletrica", "D97706"),
    ("Reboco", "1A6EE8"),
    ("Revestimento", "7C3AED"),
    ("Pintura", "D63A3A"),
    ("Esquadrias", "0EA5E9"),
    ("Loucas e Metais", "DB2777"),
    ("Limpeza Final", "65A30D"),
]
NUMERO_DO_SERVICO = {nome: i + 1 for i, (nome, _) in enumerate(SERVICOS)}

PAVIMENTOS = [f"{i:02d}PAV" for i in range(1, 9)]
PROJECT_START = datetime.date(2026, 6, 1)

OBRA = "NOME DA OBRA"  # placeholder de proposito: na ferramenta real, o engenheiro preenche
ESCOPO = "APARTAMENTOS TIPO"
DETALHE_ESCOPO = "TORRE B"
RESPONSAVEL = "Eng. Fulano de Tal"
DATA_EXPORT = datetime.date(2026, 7, 2)

# Marcos ficticios pra testar o layout: 2 propositalmente proximos (pra ver a colisao sendo
# empilhada em "andares", igual o grafico real ja faz com os badges) e 1 mais isolado.
MARCOS_FICTICIOS = [
    {"dia": 10, "nome": "Entrega Fundacao", "cor": "1A6EE8"},
    {"dia": 12, "nome": "Inicio Estrutura", "cor": "D97706"},
    {"dia": 35, "nome": "Vistoria Corpo de Bombeiros", "cor": "D63A3A"},
]

bars = []
for pi, pav in enumerate(PAVIMENTOS):
    offset = pi * 3
    cursor = offset
    for si, (nome, cor) in enumerate(SERVICOS):
        start = cursor + (si % 4) - 1
        start = max(start, offset)
        dur = 4 + (si % 3) * 3
        bars.append({"pav": pav, "servico": nome, "cor": cor, "start": start, "dur": dur})
        cursor = start + dur - 3

total_dias = max(b["start"] + b["dur"] for b in bars) + 3


def pack_lanes(bars_do_pav):
    ordenado = sorted(bars_do_pav, key=lambda b: b["start"])
    fim_por_raia = []
    for b in ordenado:
        raia = next((i for i, fim in enumerate(fim_por_raia) if fim <= b["start"]), None)
        if raia is None:
            raia = len(fim_por_raia)
            fim_por_raia.append(0)
        b["lane"] = raia
        fim_por_raia[raia] = b["start"] + b["dur"]
    return len(fim_por_raia)


num_raias_por_pav = {}
for pav in PAVIMENTOS:
    bars_do_pav = [b for b in bars if b["pav"] == pav]
    num_raias_por_pav[pav] = pack_lanes(bars_do_pav)

# ============================================================================================
# Monta a aba da linha de balanco. "simples=True" troca o texto do pacote dentro da barra por
# so o numero do servico, e reserva 3 colunas à ESQUERDA de tudo (antes da coluna de
# pavimento) pra uma legenda lateral estilizada (numero + cor + nome), ja que sem o nome
# escrito na barra precisa de uma forma de decodificar o numero.
# ============================================================================================
def construir_aba(wb, titulo_aba, simples, marcos=None):
    ws = wb.create_sheet(titulo_aba)
    marcos = marcos or []

    # Na versao simples, colunas 1-3 viram a legenda lateral e tudo o mais (pavimento +
    # calendario) desloca 3 colunas pra direita. Na versao completa, layout de sempre.
    if simples:
        COL_LEG_NUM, COL_LEG_COR, COL_LEG_NOME = 1, 2, 3
        COL_PAV = 4
    else:
        COL_PAV = 1
    COL_DIA0 = COL_PAV + 1  # primeira coluna de dia
    ultima_col = COL_DIA0 + total_dias - 1

    def linha_cabecalho(row, texto, fonte, alinhamento=None):
        ws.merge_cells(start_row=row, start_column=COL_DIA0, end_row=row, end_column=ultima_col)
        cel = ws.cell(row=row, column=COL_DIA0, value=texto)
        cel.font = fonte
        cel.alignment = alinhamento or Alignment(horizontal="left", vertical="center", indent=1)
        return cel

    def desenhar_cabecalho_calendario(row_mes, row_dia):
        limites = []
        mes_atual = None
        inicio_mes_col = COL_DIA0
        for d in range(total_dias):
            date = PROJECT_START + datetime.timedelta(days=d)
            col = COL_DIA0 + d
            mes_label = mes_label_pt(date)
            if mes_label != mes_atual:
                if mes_atual is not None:
                    ws.merge_cells(start_row=row_mes, start_column=inicio_mes_col, end_row=row_mes, end_column=col - 1)
                mes_atual = mes_label
                inicio_mes_col = col
                limites.append(col)
                cel_mes = ws.cell(row=row_mes, column=col, value=mes_label)
                cel_mes.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
                cel_mes.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_mes, column=col).fill = PatternFill("solid", fgColor=CINZA_MES_HDR)
            dow = date.weekday()
            cel_dia = ws.cell(row=row_dia, column=col, value=date.day)
            cel_dia.font = FONTE_DIA_FIMSEMANA if dow >= 5 else FONTE_DIA
            cel_dia.alignment = Alignment(horizontal="center")
            cel_dia.fill = FILL_FIMSEMANA_HDR if dow >= 5 else PatternFill("solid", fgColor=CINZA_MES_HDR)
        ws.merge_cells(start_row=row_mes, start_column=inicio_mes_col, end_row=row_mes, end_column=ultima_col)
        ws.row_dimensions[row_mes].height = 18
        ws.row_dimensions[row_dia].height = 16
        return limites

    def rotulo_pavimento(row_mes, row_dia, com_texto=True):
        ws.merge_cells(start_row=row_mes, start_column=COL_PAV, end_row=row_dia, end_column=COL_PAV)
        cel = ws.cell(row=row_mes, column=COL_PAV, value="PAVIMENTO" if com_texto else None)
        cel.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
        cel.fill = PatternFill("solid", fgColor=CINZA_PAV_COL)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = Border(left=BORDA_PAV_SEP, right=BORDA_PAV_SEP, top=BORDA_PAV_SEP, bottom=BORDA_PAV_SEP)
        ws.cell(row=row_dia, column=COL_PAV).fill = PatternFill("solid", fgColor=CINZA_PAV_COL)
        borda_com(ws.cell(row=row_dia, column=COL_PAV), left=BORDA_PAV_SEP, right=BORDA_PAV_SEP, bottom=BORDA_PAV_SEP)

    # ---------- Cabecalho: banner, tarja de contraste, subtitulo, metadados isolados ----------
    ROW_BANNER, ROW_ACCENT, ROW_SUBTITULO = 1, 2, 3
    ROW_META_RESP, ROW_META_STATS, ROW_META_DATA, ROW_SPACER = 4, 5, 6, 7

    ws.cell(row=ROW_BANNER, column=COL_PAV).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=ROW_ACCENT, column=COL_PAV).fill = PatternFill("solid", fgColor=AZUL_CONTRASTE)
    for r in range(ROW_SUBTITULO, ROW_SPACER + 1):
        ws.cell(row=r, column=COL_PAV).fill = PatternFill(fill_type=None)

    cel_banner = linha_cabecalho(ROW_BANNER, f"LINHA DE BALANCO - {OBRA}",
                                  Font(name="Calibri", bold=True, size=13, color="FFFFFF"))
    cel_banner.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[ROW_BANNER].height = 30

    ws.merge_cells(start_row=ROW_ACCENT, start_column=COL_DIA0, end_row=ROW_ACCENT, end_column=ultima_col)
    ws.cell(row=ROW_ACCENT, column=COL_DIA0).fill = PatternFill("solid", fgColor=AZUL_CONTRASTE)
    ws.row_dimensions[ROW_ACCENT].height = 4

    subtitulo_txt = f"{ESCOPO} ({DETALHE_ESCOPO})" + ("  -  versao simplificada" if simples else "")
    linha_cabecalho(ROW_SUBTITULO, subtitulo_txt, Font(name="Calibri", bold=True, size=14, color=NAVY))
    ws.row_dimensions[ROW_SUBTITULO].height = 22

    linha_cabecalho(ROW_META_RESP, f"Responsavel: {RESPONSAVEL}", Font(name="Calibri", size=10, color=MUTED))
    linha_cabecalho(ROW_META_STATS, f"{len(bars)} tarefas  |  {len(PAVIMENTOS)} pavimentos  |  {len(SERVICOS)} servicos",
                    Font(name="Calibri", size=10, color=MUTED))
    linha_cabecalho(ROW_META_DATA, f"Exportado em {DATA_EXPORT.strftime('%d/%m/%Y')}", Font(name="Calibri", size=10, color=MUTED))
    for r in range(ROW_META_RESP, ROW_META_DATA + 1):
        ws.row_dimensions[r].height = 16

    for row in range(ROW_META_RESP, ROW_META_DATA + 1):
        borda_com(ws.cell(row=row, column=COL_DIA0), left=BORDA_PAV_SEP)
        borda_com(ws.cell(row=row, column=ultima_col), right=BORDA_PAV_SEP)
    for col in range(COL_DIA0, ultima_col + 1):
        borda_com(ws.cell(row=ROW_META_RESP, column=col), top=BORDA_PAV_SEP)
        borda_com(ws.cell(row=ROW_META_DATA, column=col), bottom=BORDA_PAV_SEP)

    ws.row_dimensions[ROW_SPACER].height = 8
    proxima_linha = ROW_SPACER + 1

    # ---------- Marcos: calcula quantos "andares" de caixa sao necessarios (mesma ideia de
    # empacotamento das raias/badges: so sobe de andar quando duas caixas colidem), e reserva
    # essas linhas ACIMA do cabecalho do calendario, antes de saber onde ROW_MES cai. ----------
    def calcular_andares_marcos():
        itens = []
        for m in marcos:
            largura_cols = max(3, math.ceil(len(m["nome"]) * 0.45))
            col_centro = COL_DIA0 + m["dia"]
            esquerda = col_centro - largura_cols // 2
            direita = esquerda + largura_cols - 1
            itens.append({**m, "col_dia": col_centro, "esquerda": esquerda, "direita": direita})
        itens.sort(key=lambda it: it["esquerda"])
        direita_por_andar = []
        for it in itens:
            andar = next((i for i, d in enumerate(direita_por_andar) if d < it["esquerda"]), None)
            if andar is None:
                andar = len(direita_por_andar)
                direita_por_andar.append(0)
            it["andar"] = andar
            direita_por_andar[andar] = it["direita"]
        return itens, len(direita_por_andar)

    marcos_itens, num_andares_marcos = calcular_andares_marcos() if marcos else ([], 0)
    row_marcos0 = proxima_linha
    for i in range(num_andares_marcos):
        ws.row_dimensions[row_marcos0 + i].height = 16

    ROW_MES, ROW_DIA = row_marcos0 + num_andares_marcos, row_marcos0 + num_andares_marcos + 1
    LINHA0 = ROW_DIA + 1

    limites_mes = desenhar_cabecalho_calendario(ROW_MES, ROW_DIA)
    rotulo_pavimento(ROW_MES, ROW_DIA)

    # ---------- So na versao simples: legenda lateral (badge numero+cor + nome do servico), à
    # esquerda da coluna de pavimento. O titulo usa a MESMA formatacao do rotulo "PAVIMENTO"
    # (mesclado nas 2 linhas do cabecalho do calendario, mesmo fundo/borda/fonte), e o bloco
    # inteiro (titulo + linhas de servico) fica emoldurado por uma borda ao redor. ----------
    if simples:
        ws.merge_cells(start_row=ROW_MES, start_column=COL_LEG_NUM, end_row=ROW_DIA, end_column=COL_LEG_NOME)
        cel_leg_titulo = ws.cell(row=ROW_MES, column=COL_LEG_NUM, value="LEGENDA DOS SERVICOS")
        cel_leg_titulo.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
        cel_leg_titulo.alignment = Alignment(horizontal="center", vertical="center")
        # O merge cobre 2 linhas x 3 colunas: preenche e emoldura CADA celula do bloco (nao so
        # a de cima-esquerda), senao o fundo/borda so aparece numa fatia do titulo mesclado.
        for col in range(COL_LEG_NUM, COL_LEG_NOME + 1):
            for row in (ROW_MES, ROW_DIA):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=CINZA_PAV_COL)
            borda_com(ws.cell(row=ROW_MES, column=col), top=BORDA_PAV_SEP)
            borda_com(ws.cell(row=ROW_DIA, column=col), bottom=BORDA_PAV_SEP)
        for row in (ROW_MES, ROW_DIA):
            borda_com(ws.cell(row=row, column=COL_LEG_NUM), left=BORDA_PAV_SEP)
            borda_com(ws.cell(row=row, column=COL_LEG_NOME), right=BORDA_PAV_SEP)

        row_legenda0 = ROW_DIA + 1
        for i, (nome, cor) in enumerate(SERVICOS):
            r = row_legenda0 + i
            ws.merge_cells(start_row=r, start_column=COL_LEG_NUM, end_row=r, end_column=COL_LEG_COR)
            cel_badge = ws.cell(row=r, column=COL_LEG_NUM, value=NUMERO_DO_SERVICO[nome])
            cel_badge.fill = PatternFill("solid", fgColor=cor)
            cel_badge.font = Font(name="Calibri", bold=True, size=10, color=texto_contraste(cor))
            cel_badge.alignment = Alignment(horizontal="center", vertical="center")
            cel_nome = ws.cell(row=r, column=COL_LEG_NOME, value=nome)
            cel_nome.font = Font(name="Calibri", bold=True, size=9, color=NAVY)  # mais escuro, melhor contraste
            cel_nome.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 18

        linha_final_legenda = row_legenda0 + len(SERVICOS) - 1

        # Borda em volta de cada linha de servico (topo/base/esquerda/direita), pra parecer
        # uma tabelinha propria isolada em vez de so um contorno externo solto.
        for r in range(row_legenda0, linha_final_legenda + 1):
            borda_com(ws.cell(row=r, column=COL_LEG_NUM), top=BORDA_PAV_SEP, bottom=BORDA_PAV_SEP, left=BORDA_PAV_SEP)
            borda_com(ws.cell(row=r, column=COL_LEG_COR), top=BORDA_PAV_SEP, bottom=BORDA_PAV_SEP)
            borda_com(ws.cell(row=r, column=COL_LEG_NOME), top=BORDA_PAV_SEP, bottom=BORDA_PAV_SEP, right=BORDA_PAV_SEP)

        ws.column_dimensions[get_column_letter(COL_LEG_NUM)].width = 5
        ws.column_dimensions[get_column_letter(COL_LEG_COR)].width = 5
        ws.column_dimensions[get_column_letter(COL_LEG_NOME)].width = 16

    # ---------- Pavimentos e raias ----------
    linha_inicial_pav = {}
    linha_atual = LINHA0
    for pav in PAVIMENTOS:
        n = num_raias_por_pav[pav]
        linha_inicial_pav[pav] = linha_atual
        if n > 1:
            ws.merge_cells(start_row=linha_atual, start_column=COL_PAV, end_row=linha_atual + n - 1, end_column=COL_PAV)
        cel_pav = ws.cell(row=linha_atual, column=COL_PAV, value=pav)
        cel_pav.font = FONTE_PAV
        cel_pav.fill = PatternFill("solid", fgColor=CINZA_PAV_COL)
        cel_pav.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # Borda em volta da celula do pavimento (mesclada por N raias): topo/base so nas
        # pontas do bloco, esquerda/direita em toda a extensao mesclada.
        for r in range(linha_atual, linha_atual + n):
            ws.cell(row=r, column=COL_PAV).fill = PatternFill("solid", fgColor=CINZA_PAV_COL)
            borda_com(ws.cell(row=r, column=COL_PAV), left=BORDA_PAV_SEP, right=BORDA_PAV_SEP)
        borda_com(ws.cell(row=linha_atual, column=COL_PAV), top=BORDA_PAV_SEP)
        borda_com(ws.cell(row=linha_atual + n - 1, column=COL_PAV), bottom=BORDA_PAV_SEP)
        linha_atual += n
    total_linhas = linha_atual - 1

    for pi, pav in enumerate(PAVIMENTOS):
        linha0 = linha_inicial_pav[pav]
        n = num_raias_por_pav[pav]
        fill_linha = FILL_PAV_PAR_O if pi % 2 == 0 else PatternFill(fill_type=None)
        for r in range(linha0, linha0 + n):
            for d in range(total_dias):
                col = COL_DIA0 + d
                date = PROJECT_START + datetime.timedelta(days=d)
                cel = ws.cell(row=r, column=col)
                if date.weekday() >= 5:
                    cel.fill = FILL_FIMSEMANA
                elif pi % 2 == 0:
                    cel.fill = fill_linha

    # ---------- Barras ----------
    for b in bars:
        row = linha_inicial_pav[b["pav"]] + b["lane"]
        fill = PatternFill("solid", fgColor=b["cor"])
        col_ini = COL_DIA0 + b["start"]
        col_fim = COL_DIA0 + b["start"] + b["dur"] - 1
        for col in range(col_ini, col_fim + 1):
            cel = ws.cell(row=row, column=col)
            cel.fill = fill
            cel.border = BORDA_FINA
        if col_fim > col_ini:
            ws.merge_cells(start_row=row, start_column=col_ini, end_row=row, end_column=col_fim)
        if simples:
            celula_texto = ws.cell(row=row, column=col_ini, value=NUMERO_DO_SERVICO[b["servico"]])
            celula_texto.font = Font(name="Calibri", bold=True, size=10, color=texto_contraste(b["cor"]))
            celula_texto.alignment = Alignment(horizontal="center", vertical="center")
        else:
            celula_texto = ws.cell(row=row, column=col_ini, value=f" {b['servico']} ({b['dur']} dias)")
            celula_texto.font = Font(name="Calibri", bold=True, size=8, color=texto_contraste(b["cor"]))
            celula_texto.alignment = Alignment(horizontal="left", vertical="center")

    # ---------- Separadores ----------
    for pav in PAVIMENTOS:
        ultima_linha_bloco = linha_inicial_pav[pav] + num_raias_por_pav[pav] - 1
        for col in range(COL_PAV, ultima_col + 1):
            borda_com(ws.cell(row=ultima_linha_bloco, column=col), bottom=BORDA_PAV_SEP)

    for col in range(COL_PAV, ultima_col + 1):
        borda_com(ws.cell(row=ROW_DIA, column=col), bottom=BORDA_PAV_SEP)

    # ---------- Rodape: repete a regua de datas ----------
    ROW_MES_RODAPE = total_linhas + 1
    ROW_DIA_RODAPE = total_linhas + 2
    desenhar_cabecalho_calendario(ROW_MES_RODAPE, ROW_DIA_RODAPE)
    rotulo_pavimento(ROW_MES_RODAPE, ROW_DIA_RODAPE, com_texto=False)
    for col in range(COL_PAV, ultima_col + 1):
        borda_com(ws.cell(row=ROW_MES_RODAPE, column=col), top=BORDA_PAV_SEP)
        borda_com(ws.cell(row=ROW_DIA_RODAPE, column=col), bottom=BORDA_PAV_SEP)
    ultima_linha_planilha = ROW_DIA_RODAPE

    for row in range(ROW_BANNER, ultima_linha_planilha + 1):
        borda_com(ws.cell(row=row, column=COL_PAV), right=BORDA_PAV_SEP)
        borda_com(ws.cell(row=row, column=ultima_col), right=BORDA_PAV_SEP)

    # Limite de mes: divisor vertical entre meses + moldura horizontal (topo/base) na propria
    # linha do rotulo do mes, no cabecalho e no rodape, pra fechar a caixa dos 4 lados.
    for col in limites_mes:
        for row in range(ROW_MES, ultima_linha_planilha + 1):
            borda_com(ws.cell(row=row, column=col), left=BORDA_MES)
    for row_mes_alvo in (ROW_MES, ROW_MES_RODAPE):
        for col in range(COL_DIA0, ultima_col + 1):
            borda_com(ws.cell(row=row_mes_alvo, column=col), top=BORDA_MES, bottom=BORDA_MES)

    # Borda vertical separando cada dia (mais forte no cabecalho/rodape, mais clara no corpo)
    linhas_cabecalho_calendario = {ROW_MES, ROW_DIA, ROW_MES_RODAPE, ROW_DIA_RODAPE}
    for d in range(total_dias):
        col = COL_DIA0 + d
        if col in limites_mes:
            continue
        for row in range(ROW_MES, ultima_linha_planilha + 1):
            borda = BORDA_DIA_HDR if row in linhas_cabecalho_calendario else BORDA_DIA_CLARA
            borda_com(ws.cell(row=row, column=col), left=borda)

    # ---------- Marcos: caixa com o nome (cor do marco, mesclada) no andar calculado antes, e
    # uma linha pontilhada na cor do marco descendo até o fim da tabela. Roda por ÚLTIMO, depois
    # de toda borda de mes/dia/pavimento, senao esses trechos passam por cima e apagam a linha
    # (foi exatamente o bug visto: o separador fino de cada dia rodava depois e sobrescrevia).
    # Limitacao real do Excel que continua existindo mesmo assim: quando o dia do marco cai
    # DENTRO de uma barra mesclada, a borda nao aparece ali (uma celula mesclada so tem borda
    # nas bordas externas dela, nunca no meio) - a linha fica com uma lacuna atras da barra
    # nesse trecho, e volta a aparecer assim que a barra termina.
    for it in marcos_itens:
        row_label = row_marcos0 + it["andar"]
        c_ini, c_fim = max(it["esquerda"], COL_DIA0), min(it["direita"], ultima_col)
        if c_fim > c_ini:
            ws.merge_cells(start_row=row_label, start_column=c_ini, end_row=row_label, end_column=c_fim)
        cel = ws.cell(row=row_label, column=c_ini, value=it["nome"])
        cel.fill = PatternFill("solid", fgColor=it["cor"])
        cel.font = Font(name="Calibri", bold=True, size=8, color=texto_contraste(it["cor"]))
        cel.alignment = Alignment(horizontal="center", vertical="center")

        borda_pontilhada = Side(style="mediumDashed", color=it["cor"])
        for r in range(row_label, ultima_linha_planilha + 1):
            borda_com(ws.cell(row=r, column=it["col_dia"]), left=borda_pontilhada)

    ws.column_dimensions[get_column_letter(COL_PAV)].width = 14
    for d in range(total_dias):
        ws.column_dimensions[get_column_letter(COL_DIA0 + d)].width = 3.00
    for r in range(LINHA0, ultima_linha_planilha + 1):
        ws.row_dimensions[r].height = 15
    ws.row_dimensions[ROW_DIA].height = 16
    ws.row_dimensions[ROW_DIA_RODAPE].height = 16
    ws.freeze_panes = ws.cell(row=LINHA0, column=COL_DIA0)
    ws.sheet_view.showGridLines = False
    return ws


wb = Workbook()
wb.remove(wb.active)  # a aba padrao "Sheet" e substituida pelas duas variacoes abaixo

# ============================ ABA 1: LINHA DE BALANCO (completa, com texto) ============================
construir_aba(wb, "Linha de Balanco", simples=False, marcos=MARCOS_FICTICIOS)

# ============================ ABA 2: LINHA DE BALANCO SIMPLES (numeros + legenda) ============================
construir_aba(wb, "Linha de Balanco Simples", simples=True, marcos=MARCOS_FICTICIOS)

# ============================ ABA 3: LEGENDA ============================
wl = wb.create_sheet("Legenda")
for i, titulo in enumerate(["NUMERO", "SERVICO", "COR"]):
    c = wl.cell(row=1, column=1 + i, value=titulo)
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
for i, (nome, cor) in enumerate(SERVICOS):
    r = 2 + i
    wl.cell(row=r, column=1, value=i + 1).font = Font(name="Calibri", bold=True, size=10)
    wl.cell(row=r, column=2, value=nome).font = Font(name="Calibri", size=10)
    swatch = wl.cell(row=r, column=3, value="")
    swatch.fill = PatternFill("solid", fgColor=cor)
wl.column_dimensions["A"].width = 10
wl.column_dimensions["B"].width = 22
wl.column_dimensions["C"].width = 10

# ============================ ABA 4: DADOS (tabela crua) ============================
wd = wb.create_sheet("Dados")
cabecalho = ["Pavimento", "Servico", "Inicio", "Fim", "Duracao (dias)", "Raia"]
for i, h in enumerate(cabecalho):
    c = wd.cell(row=1, column=1 + i, value=h)
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
for i, b in enumerate(sorted(bars, key=lambda x: (x["pav"], x["lane"], x["start"]))):
    r = 2 + i
    data_ini = PROJECT_START + datetime.timedelta(days=b["start"])
    data_fim = PROJECT_START + datetime.timedelta(days=b["start"] + b["dur"] - 1)
    wd.cell(row=r, column=1, value=b["pav"])
    wd.cell(row=r, column=2, value=b["servico"])
    wd.cell(row=r, column=3, value=data_ini).number_format = "DD/MM/YYYY"
    wd.cell(row=r, column=4, value=data_fim).number_format = "DD/MM/YYYY"
    wd.cell(row=r, column=5, value=b["dur"])
    wd.cell(row=r, column=6, value=b["lane"] + 1)
for col, w in zip("ABCDEF", [12, 18, 12, 12, 14, 8]):
    wd.column_dimensions[col].width = w

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "export-excel-preview.xlsx")
wb.save(OUT)
print("Salvo em:", OUT)
print("Raias por pavimento:", num_raias_por_pav)
